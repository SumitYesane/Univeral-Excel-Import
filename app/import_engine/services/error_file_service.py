import os
from collections import defaultdict
from copy import deepcopy

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.core.config import settings

SOURCE_ROW_FIELD = "__source_row__"
SOURCE_SHEET_FIELD = "__source_sheet__"
SOURCE_MODEL_FIELD = "__source_model__"
ERROR_MESSAGE_FIELD = "__error__"
ERROR_COLUMN_NAME = "Error"

HEADER_FILL = PatternFill(fill_type="solid", fgColor="203040")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ERROR_FILL = PatternFill(fill_type="solid", fgColor="FDE8E8")
ERROR_CELL_FILL = PatternFill(fill_type="solid", fgColor="F9C0C0")


def _strip_metadata(data: dict) -> dict:
    return {
        key: value
        for key, value in data.items()
        if key not in {SOURCE_ROW_FIELD, SOURCE_SHEET_FIELD, SOURCE_MODEL_FIELD, ERROR_MESSAGE_FIELD}
    }


def build_error_entry(row, sheet, model_name: str, data: dict, error_message: str) -> dict:
    payload = deepcopy(data)
    sheet_name = sheet or payload.get(SOURCE_SHEET_FIELD) or "unknown_sheet"
    source_row = row or payload.get(SOURCE_ROW_FIELD)
    payload[SOURCE_SHEET_FIELD] = sheet_name
    payload[SOURCE_ROW_FIELD] = source_row
    payload[SOURCE_MODEL_FIELD] = model_name
    return {
        "sheet": sheet_name,
        "row": source_row,
        "model": model_name,
        "data": payload,
        "error": error_message,
    }


def _sheet_title(sheet_name: str) -> str:
    invalid = set('[]:*?/\\')
    cleaned = "".join("_" if char in invalid else char for char in str(sheet_name))
    return (cleaned or "sheet")[:31]


def _group_errors(errors):
    grouped = defaultdict(list)
    for entry in errors:
        grouped[entry.get("sheet") or "unknown_sheet"].append(entry)
    return grouped


def _write_rows(worksheet, rows):
    if not rows:
        worksheet.append(["No errors"])
        return

    data_columns = []
    for row in rows:
        for key in _strip_metadata(row["data"]).keys():
            if key not in data_columns:
                data_columns.append(key)

    headers = data_columns + [ERROR_COLUMN_NAME]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for entry in rows:
        clean_data = _strip_metadata(entry["data"])
        worksheet.append([clean_data.get(col) for col in data_columns] + [entry.get("error")])
        current_row = worksheet.max_row
        for cell in worksheet[current_row]:
            cell.fill = ERROR_FILL
        worksheet.cell(current_row, len(headers)).fill = ERROR_CELL_FILL

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def build_error_file(errors, job_id):
    settings.ERROR_ROOT.mkdir(parents=True, exist_ok=True)
    path = os.path.join(settings.ERROR_ROOT, f"{job_id}_errors.xlsx")
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"

    grouped = _group_errors(errors)
    summary_rows = []
    for sheet_name, entries in grouped.items():
        summary_rows.append(
            {
                "sheet": sheet_name,
                "error_count": len(entries),
                "source_rows": ", ".join(str(entry.get("row")) for entry in entries if entry.get("row") is not None),
            }
        )

    _write_rows(
        summary,
        [
            {
                "row": row["source_rows"],
                "model": row["sheet"],
                "error": f'{row["error_count"]} error(s)',
                "data": {"Sheet": row["sheet"], "Rows": row["source_rows"]},
            }
            for row in summary_rows
        ],
    )

    for sheet_name, entries in grouped.items():
        worksheet = workbook.create_sheet(title=_sheet_title(sheet_name))
        _write_rows(worksheet, entries)

    workbook.save(path)
    return path
